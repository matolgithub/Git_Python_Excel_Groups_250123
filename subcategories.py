import openpyxl as op
from pprint import pprint as pp
from datetime import datetime as dt


def get_data(filename="order_table_250123.xlsx"):
    subcategories_dict = {}

    wb = op.load_workbook(filename=filename, data_only=True)
    sheet = wb.active

    max_rows = sheet.max_row  # 748

    for item in range(7, max_rows + 1):
        sku = sheet.cell(row=item, column=2).value
        subcategory = sheet.cell(row=item, column=12).value

        if not sku:
            continue

        if subcategory not in subcategories_dict:
            subcategories_dict[subcategory] = [sku]
        else:
            subcategories_dict[subcategory].append(sku)

    sorted_dict = dict(sorted(subcategories_dict.items()))
    pp(sorted_dict)

    return sorted_dict


def write_file(file_name="subcategories.ini"):
    subcategories = get_data()
    start_time = dt.now()

    with open(file=file_name, mode="w") as my_newfile:
        for key, value in subcategories.items():
            string_values = ", ".join(value)
            string_to_write = f"{key}: {string_values}\n"
            my_newfile.write(string_to_write)

    total_time = dt.now() - start_time

    print(f"The file '{file_name}' was created at: {dt.now()}. Total operation time is: {total_time}.")


if __name__ == "__main__":
    write_file()
